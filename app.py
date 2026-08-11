import asyncio
import io
import json
import logging
import os
import re
import subprocess
import sys
import time
import uuid
import zipfile
from pathlib import Path
from typing import Any, AsyncGenerator

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import HTMLResponse, StreamingResponse, Response
from fastapi.staticfiles import StaticFiles
from openai import AsyncOpenAI
from pydantic import BaseModel, Field


# ============================================================
# CONFIG
# ============================================================

load_dotenv()

APP_NAME = os.getenv("APP_NAME", "AI Canvas")

BASE_URL = (
    os.getenv(
        "BASE_URL",
        "https://freemodelsforall.hopto.org/v1",
    )
    .rstrip("/")
)

API_KEY = (
    os.getenv("API_KEY")
    or os.getenv("OPENAI_API_KEY")
)

DEFAULT_MODEL = os.getenv(
    "DEFAULT_MODEL",
    "obsidianx/openai/gpt-5-6-luna",
)

WORKSPACE_ROOT = Path(
    os.getenv(
        "WORKSPACE_ROOT",
        "./workspace",
    )
).resolve()

MAX_FILE_BYTES = int(
    os.getenv(
        "MAX_FILE_BYTES",
        str(25 * 1024 * 1024),
    )
)

MAX_TOOL_ROUNDS = int(
    os.getenv(
        "MAX_TOOL_ROUNDS",
        "16",
    )
)

CODE_TIMEOUT = int(
    os.getenv(
        "CODE_TIMEOUT_SECONDS",
        "20",
    )
)

ENABLE_CODE_EXECUTION = (
    os.getenv(
        "ENABLE_CODE_EXECUTION",
        "true",
    ).lower()
    == "true"
)

AI_TIMEOUT = float(
    os.getenv(
        "AI_TIMEOUT_SECONDS",
        "180",
    )
)

AI_RETRIES = int(
    os.getenv(
        "AI_RETRIES",
        "3",
    )
)

WORKSPACE_ROOT.mkdir(
    parents=True,
    exist_ok=True,
)


# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format=(
        "%(asctime)s | "
        "%(levelname)s | "
        "%(name)s | "
        "%(message)s"
    ),
)

logger = logging.getLogger(
    "ai_canvas"
)


# ============================================================
# FALLBACK MODELS
# ============================================================

FALLBACK_MODELS = [
    (
        "Claude Fable 5",
        "obsidianx/custom/claude-fable-5",
    ),
    (
        "Gpt 5 4",
        "obsidianx/openai/gpt-5-4",
    ),
    (
        "Gpt 5 5",
        "obsidianx/openai/gpt-5-5",
    ),
    (
        "Gpt 5 6 Luna",
        "obsidianx/openai/gpt-5-6-luna",
    ),
    (
        "Gpt 5 6 Terra",
        "obsidianx/openai/gpt-5-6-terra",
    ),
    (
        "Gpt 5.6",
        "obsidianx/openai/gpt-5.6",
    ),
    (
        "Claude Opus 5",
        "obsidianx/custom/claude-opus-5",
    ),
    (
        "Claude Sonnet 5",
        "obsidianx/custom/claude-sonnet-5",
    ),
    (
        "DeepSeek V4 Pro",
        "obsidianx/openai/deepseek-v4-pro",
    ),
    (
        "Gemini 3 Pro",
        "obsidianx/custom/gemini-3-pro",
    ),
    (
        "Grok 4 5",
        "obsidianx/custom/grok-4-5",
    ),
    (
        "Longcat 2.0 Free",
        "obsidianx/custom/longcat-2.0-free",
    ),
    (
        "North Mini Code Free",
        "obsidianx/custom/north-mini-code-free",
    ),
]


# ============================================================
# AGENT SYSTEM PROMPT
# ============================================================

AGENT_SYSTEM_PROMPT = """
You are the reasoning and execution engine inside AI Canvas.

Behave like a careful senior human expert and an autonomous
task agent, not a one-shot chatbot.

CORE BEHAVIOR

- Understand the user's actual objective before acting.
- For complex requests, silently create a practical plan.
- Execute the task step by step.
- Use tools whenever they materially improve correctness.
- Inspect files before making claims about them.
- Verify calculations, code, transformations and outputs.
- Never claim that you performed an action unless the tool result
  proves it.
- After every tool action, inspect the result and decide what
  should happen next.
- Continue working until the requested outcome is actually
  complete or a genuine external limitation blocks it.
- If something is uncertain, state the uncertainty honestly.
- Never invent facts, files, sources, citations or results.
- Never promise 100% accuracy.
- Maximize accuracy through verification.

RESPONSE QUALITY

- Produce complete sentences.
- Produce coherent connected answers.
- Never output broken fragments.
- Never expose internal tool syntax.
- Never expose hidden reasoning.
- Never output fake streaming artifacts.
- Do not repeat the user's prompt unnecessarily.
- Avoid generic filler.
- Match the user's language when practical.
- Sound natural and human.
- Use headings and bullets when useful.
- Be concise for simple questions and detailed for complex work.

FILE WORK

- Identify relevant uploaded files first.
- Inspect/extract files before making claims about them.
- Preserve existing structure when editing.
- Re-check important modifications.
- Intermediate files belong in files/.
- Final user-facing deliverables belong in outputs/.
- Never delete or overwrite important user files unless clearly
  requested.

AGENT LOOP

Plan
→ inspect
→ act
→ observe
→ verify
→ fix/retry
→ deliver.

Stop only when the requested outcome is complete, a necessary
capability is genuinely unavailable, or user approval is required.
"""


# ============================================================
# TOOLS
# ============================================================

TOOLS_SCHEMA = [
    {
        "type": "function",
        "function": {
            "name": "write_file",
            "description": (
                "Create or overwrite a UTF-8 working/source file "
                "inside the current chat workspace."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string"
                    },
                    "content": {
                        "type": "string"
                    },
                },
                "required": [
                    "filepath",
                    "content",
                ],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_output_file",
            "description": (
                "Create a final downloadable deliverable "
                "inside outputs/."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filename": {
                        "type": "string"
                    },
                    "content": {
                        "type": "string"
                    },
                },
                "required": [
                    "filename",
                    "content",
                ],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_file",
            "description": (
                "Delete a file inside the current workspace."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string"
                    }
                },
                "required": [
                    "filepath"
                ],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": (
                "Read a file from the current workspace."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string"
                    }
                },
                "required": [
                    "filepath"
                ],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "search_files",
            "description": (
                "Search filenames and text-readable files."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string"
                    },
                    "directory": {
                        "type": "string",
                        "default": ".",
                    },
                },
                "required": [
                    "query"
                ],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_directory",
            "description": (
                "List files and folders in the workspace."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "directory": {
                        "type": "string",
                        "default": ".",
                    }
                },
                "required": [],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "run_python",
            "description": (
                "Run a short Python program for calculations, "
                "data processing, validation or file generation."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string"
                    },
                    "timeout": {
                        "type": "integer",
                        "minimum": 1,
                        "maximum": 60,
                    },
                },
                "required": [
                    "code"
                ],
            },
        },
    },
]


# ============================================================
# APP
# ============================================================

app = FastAPI(
    title=APP_NAME,
    version="4.1.0",
)

app.mount(
    "/static",
    StaticFiles(
        directory="static"
    ),
    name="static",
)


# ============================================================
# MODELS
# ============================================================

class ChatRequest(BaseModel):
    chat_id: str = Field(
        default="default",
        min_length=1,
        max_length=100,
    )

    model: str = Field(
        default=DEFAULT_MODEL,
        min_length=1,
        max_length=200,
    )

    messages: list[
        dict[str, Any]
    ] = Field(
        default_factory=list
    )

    tools_enabled: bool = True


class DownloadRequest(BaseModel):
    chat_id: str = Field(
        min_length=1,
        max_length=100,
    )

    title: str = Field(
        default="AI Canvas Chat",
        max_length=200,
    )

    messages: list[
        dict[str, Any]
    ] = Field(
        default_factory=list
    )


# ============================================================
# SSE
# ============================================================

def sse(
    event: str,
    data: Any,
) -> str:
    return (
        f"event: {event}\n"
        f"data: "
        f"{json.dumps(data, ensure_ascii=False)}"
        f"\n\n"
    )


# ============================================================
# WORKSPACE
# ============================================================

def chat_root(
    chat_id: str
) -> Path:

    if not re.fullmatch(
        r"[A-Za-z0-9_-]{1,100}",
        chat_id,
    ):
        raise ValueError(
            "Invalid chat id"
        )

    root = (
        WORKSPACE_ROOT
        / "chats"
        / chat_id
    ).resolve()

    if (
        WORKSPACE_ROOT
        not in root.parents
    ):
        raise ValueError(
            "Invalid chat workspace"
        )

    for directory in (
        "uploads",
        "outputs",
        "files",
    ):
        (
            root / directory
        ).mkdir(
            parents=True,
            exist_ok=True,
        )

    return root


def safe_path(
    root: Path,
    relative: str,
) -> Path:

    relative = (
        relative or ""
    ).strip().replace(
        "\\",
        "/",
    )

    if not relative:
        raise ValueError(
            "A relative path is required."
        )

    candidate = (
        root / relative
    ).resolve()

    if (
        candidate != root
        and root not in candidate.parents
    ):
        raise ValueError(
            "Path escapes the workspace."
        )

    return candidate


# ============================================================
# FILE EXTRACTION
# ============================================================

def extract_file(
    path: Path
) -> str:

    ext = path.suffix.lower()

    text_exts = {
        ".txt",
        ".md",
        ".py",
        ".js",
        ".ts",
        ".tsx",
        ".jsx",
        ".html",
        ".css",
        ".json",
        ".csv",
        ".xml",
        ".yaml",
        ".yml",
        ".toml",
        ".ini",
        ".log",
        ".sql",
        ".sh",
        ".c",
        ".cpp",
        ".java",
        ".rs",
        ".go",
    }

    if ext in text_exts:
        return path.read_text(
            encoding="utf-8",
            errors="replace",
        )

    if ext == ".pdf":
        from pypdf import PdfReader

        reader = PdfReader(
            str(path)
        )

        return "\n\n".join(
            (
                page.extract_text()
                or ""
            )
            for page in reader.pages
        )

    if ext == ".docx":
        from docx import Document

        document = Document(
            str(path)
        )

        return "\n".join(
            paragraph.text
            for paragraph
            in document.paragraphs
        )

    if ext == ".xlsx":
        from openpyxl import (
            load_workbook
        )

        workbook = load_workbook(
            str(path),
            read_only=True,
            data_only=True,
        )

        chunks = []

        for sheet in workbook.worksheets:
            chunks.append(
                f"[Sheet: {sheet.title}]"
            )

            for row in sheet.iter_rows(
                values_only=True
            ):
                chunks.append(
                    "\t".join(
                        ""
                        if value is None
                        else str(value)
                        for value in row
                    )
                )

        return "\n".join(
            chunks
        )

    raise ValueError(
        "Unsupported text extraction format: "
        f"{ext or 'unknown'}"
    )


# ============================================================
# TOOL EXECUTION
# ============================================================

def execute_tool(
    name: str,
    arguments: str,
    chat_id: str,
) -> str:

    try:
        try:
            args = (
                json.loads(arguments)
                if arguments.strip()
                else {}
            )
        except json.JSONDecodeError as exc:
            return (
                "Error: invalid tool "
                f"arguments JSON: {exc}"
            )

        root = chat_root(
            chat_id
        )

        if name == "write_file":

            path = safe_path(
                root,
                args.get(
                    "filepath",
                    "",
                ),
            )

            content = str(
                args.get(
                    "content",
                    "",
                )
            )

            if (
                len(
                    content.encode()
                )
                > MAX_FILE_BYTES
            ):
                return (
                    "Error: file exceeds "
                    "size limit."
                )

            path.parent.mkdir(
                parents=True,
                exist_ok=True,
            )

            path.write_text(
                content,
                encoding="utf-8",
            )

            return (
                "Success: wrote "
                f"'{path.relative_to(root)}'."
            )

        if name == "write_output_file":

            filename = Path(
                str(
                    args.get(
                        "filename",
                        "output.txt",
                    )
                )
            ).name

            if (
                not filename
                or filename.startswith(".")
            ):
                filename = "output.txt"

            path = (
                root
                / "outputs"
                / filename
            )

            content = str(
                args.get(
                    "content",
                    "",
                )
            )

            if (
                len(
                    content.encode()
                )
                > MAX_FILE_BYTES
            ):
                return (
                    "Error: output exceeds "
                    "size limit."
                )

            path.write_text(
                content,
                encoding="utf-8",
            )

            return (
                "Success: final output "
                f"saved as "
                f"'outputs/{filename}'."
            )

        if name == "delete_file":

            path = safe_path(
                root,
                args.get(
                    "filepath",
                    "",
                ),
            )

            if (
                not path.exists()
                or not path.is_file()
            ):
                return (
                    "Error: file not found."
                )

            path.unlink()

            return (
                "Success: deleted "
                f"'{path.relative_to(root)}'."
            )

        if name == "read_file":

            path = safe_path(
                root,
                args.get(
                    "filepath",
                    "",
                ),
            )

            if not path.is_file():
                return (
                    "Error: file not found."
                )

            if (
                path.stat().st_size
                > MAX_FILE_BYTES
            ):
                return (
                    "Error: file exceeds "
                    "size limit."
                )

            try:
                return extract_file(
                    path
                )
            except Exception as exc:
                return (
                    "Error extracting "
                    f"'{path.relative_to(root)}': "
                    f"{exc}"
                )

        if name == "search_files":

            query = str(
                args.get(
                    "query",
                    "",
                )
            ).lower().strip()

            if not query:
                return (
                    "Error: query is required."
                )

            directory = args.get(
                "directory"
            )

            if directory in (
                None,
                "",
                ".",
            ):
                base = root
            else:
                base = safe_path(
                    root,
                    directory,
                )

            hits = []

            for path in base.rglob("*"):

                if not path.is_file():
                    continue

                relative = (
                    path
                    .relative_to(root)
                    .as_posix()
                )

                if (
                    query
                    in path.name.lower()
                ):
                    hits.append(
                        {
                            "file": relative,
                            "match": "filename",
                        }
                    )

                    continue

                try:

                    if (
                        path.stat().st_size
                        > MAX_FILE_BYTES
                    ):
                        continue

                    text = extract_file(
                        path
                    )

                    position = (
                        text.lower()
                        .find(query)
                    )

                    if position >= 0:
                        hits.append(
                            {
                                "file": relative,
                                "match": text[
                                    max(
                                        0,
                                        position - 120,
                                    ):
                                    position
                                    + len(query)
                                    + 240
                                ],
                            }
                        )

                except Exception:
                    pass

                if len(hits) >= 50:
                    break

            return json.dumps(
                hits,
                ensure_ascii=False,
                indent=2,
            )

        if name == "list_directory":

            directory = args.get(
                "directory"
            )

            if directory in (
                None,
                "",
                ".",
            ):
                path = root
            else:
                path = safe_path(
                    root,
                    directory,
                )

            if not path.is_dir():
                return (
                    "Error: directory not found."
                )

            items = []

            for item in sorted(
                path.iterdir(),
                key=lambda p: (
                    not p.is_dir(),
                    p.name.lower(),
                ),
            ):
                items.append(
                    {
                        "name": item.name,
                        "type": (
                            "directory"
                            if item.is_dir()
                            else "file"
                        ),
                        "size": (
                            item.stat().st_size
                            if item.is_file()
                            else None
                        ),
                    }
                )

            return json.dumps(
                items,
                ensure_ascii=False,
                indent=2,
            )

        if name == "run_python":

            if not ENABLE_CODE_EXECUTION:
                return (
                    "Error: Python execution "
                    "is disabled by server "
                    "configuration."
                )

            code = str(
                args.get(
                    "code",
                    "",
                )
            )

            timeout = min(
                max(
                    int(
                        args.get(
                            "timeout",
                            CODE_TIMEOUT,
                        )
                    ),
                    1,
                ),
                60,
            )

            proc = subprocess.run(
                [
                    sys.executable,
                    "-I",
                    "-c",
                    code,
                ],
                cwd=str(root),
                capture_output=True,
                text=True,
                timeout=timeout,
                env={
                    "PATH": os.getenv(
                        "PATH",
                        "",
                    )
                },
            )

            output = (
                proc.stdout or ""
            )

            if proc.stderr:
                output += (
                    "\n[stderr]\n"
                    + proc.stderr
                )

            return (
                f"Exit code: "
                f"{proc.returncode}\n"
                f"{output[-20000:]}"
            )

        return (
            f"Error: unknown tool "
            f"'{name}'."
        )

    except subprocess.TimeoutExpired:
        return (
            "Error: Python execution "
            "timed out."
        )

    except Exception as exc:
        logger.exception(
            "Tool execution failed: %s",
            name,
        )

        return (
            f"Error executing {name}: "
            f"{exc}"
        )


# ============================================================
# MODEL HELPERS
# ============================================================

async def get_models(
    client: AsyncOpenAI
) -> list[str]:

    try:
        response = (
            await client.models.list()
        )

        ids = sorted(
            {
                model.id
                for model in response.data
                if getattr(
                    model,
                    "id",
                    None,
                )
            }
        )

        return (
            ids
            or [
                item[1]
                for item in FALLBACK_MODELS
            ]
        )

    except Exception as exc:

        logger.warning(
            "Model listing failed: %s",
            exc,
        )

        return [
            item[1]
            for item in FALLBACK_MODELS
        ]


def build_messages(
    raw: list[
        dict[str, Any]
    ]
) -> list[
    dict[str, Any]
]:

    cleaned = []

    for message in raw:

        role = message.get(
            "role"
        )

        if role in {
            "user",
            "assistant",
            "tool",
        }:
            cleaned.append(
                message
            )

    return [
        {
            "role": "system",
            "content":
                AGENT_SYSTEM_PROMPT,
        }
    ] + cleaned


# ============================================================
# ERROR CLASSIFICATION
# ============================================================

def error_text(
    exc: Exception
) -> str:

    return (
        f"{type(exc).__name__}: "
        f"{str(exc)}"
    )


def is_transient_error(
    exc: Exception
) -> bool:

    text = str(
        exc
    ).lower()

    transient_markers = (
        "429",
        "500",
        "502",
        "503",
        "504",
        "rate limit",
        "temporarily unavailable",
        "service unavailable",
        "bad gateway",
        "gateway timeout",
        "connection reset",
        "connection refused",
        "connection error",
        "server disconnected",
        "timed out",
        "timeout",
    )

    return any(
        marker in text
        for marker in transient_markers
    )


# ============================================================
# SINGLE MODEL STREAM
# ============================================================

async def call_model_stream(
    client: AsyncOpenAI,
    model: str,
    messages: list[
        dict[str, Any]
    ],
    tools_enabled: bool,
) -> tuple[
    str,
    dict[int, dict[str, str]],
    bool,
]:

    kwargs = {
        "model": model,
        "messages": messages,
        "stream": True,
    }

    if tools_enabled:
        kwargs[
            "tools"
        ] = TOOLS_SCHEMA

    logger.info(
        "AI request started | "
        "model=%s | tools=%s",
        model,
        tools_enabled,
    )

    response = (
        await client.chat.completions.create(
            **kwargs
        )
    )

    assistant_text = ""
    tool_calls = {}
    got_any_chunk = False

    async for chunk in response:

        got_any_chunk = True

        if not chunk.choices:
            continue

        delta = (
            chunk.choices[0].delta
        )

        content = getattr(
            delta,
            "content",
            None,
        )

        if content:
            assistant_text += content

        incoming = getattr(
            delta,
            "tool_calls",
            None,
        )

        if incoming:

            for tool_call in incoming:

                index = getattr(
                    tool_call,
                    "index",
                    0,
                )

                entry = (
                    tool_calls.setdefault(
                        index,
                        {
                            "id": None,
                            "name": "",
                            "arguments": "",
                        },
                    )
                )

                if getattr(
                    tool_call,
                    "id",
                    None,
                ):
                    entry[
                        "id"
                    ] = tool_call.id

                function = getattr(
                    tool_call,
                    "function",
                    None,
                )

                if function:

                    name = getattr(
                        function,
                        "name",
                        None,
                    )

                    arguments = getattr(
                        function,
                        "arguments",
                        None,
                    )

                    if name:
                        entry[
                            "name"
                        ] = name

                    if arguments:
                        entry[
                            "arguments"
                        ] += arguments

    if (
        not got_any_chunk
        and not tool_calls
    ):
        raise RuntimeError(
            "The AI gateway returned "
            "an empty stream."
        )

    return (
        assistant_text,
        tool_calls,
        got_any_chunk,
    )


# ============================================================
# AGENT STREAM
# ============================================================

async def stream_chat(
    req: ChatRequest,
) -> AsyncGenerator[
    str,
    None,
]:

    if not API_KEY:

        yield sse(
            "error",
            {
                "message":
                    "API_KEY is not configured "
                    "on the Railway server."
            },
        )

        return

    client = AsyncOpenAI(
        base_url=BASE_URL,
        api_key=API_KEY,
        timeout=AI_TIMEOUT,
        max_retries=0,
    )

    messages = build_messages(
        req.messages
    )

    requested_model = (
        req.model.strip()
        or DEFAULT_MODEL
    )

    tools_enabled = (
        req.tools_enabled
    )

    yield sse(
        "status",
        {
            "message":
                "Agent is planning "
                "and connecting…",
            "model":
                requested_model,
        },
    )

    logger.info(
        "New chat request | "
        "chat_id=%s | model=%s | "
        "base_url=%s",
        req.chat_id,
        requested_model,
        BASE_URL,
    )

    current_model = (
        requested_model
    )

    model_attempts = [
        current_model
    ]

    for fallback in (
        "obsidianx/openai/gpt-5-6-luna",
        "obsidianx/openai/gpt-5-5",
        "obsidianx/openai/gpt-5-4",
    ):
        if fallback not in model_attempts:
            model_attempts.append(
                fallback
            )

    model_index = 0

    for round_index in range(
        MAX_TOOL_ROUNDS
    ):

        if (
            model_index
            >= len(model_attempts)
        ):
            model_index = 0

        current_model = (
            model_attempts[
                model_index
            ]
        )

        transient_failures = 0

        while (
            transient_failures
            <= AI_RETRIES
        ):

            try:

                (
                    assistant_text,
                    tool_calls,
                    _,
                ) = await asyncio.wait_for(
                    call_model_stream(
                        client,
                        current_model,
                        messages,
                        tools_enabled,
                    ),
                    timeout=AI_TIMEOUT,
                )

                # ------------------------------------------------
                # If the gateway accepted the request but returned
                # no text and no tools, don't silently finish.
                # ------------------------------------------------

                if (
                    not assistant_text.strip()
                    and not tool_calls
                ):
                    raise RuntimeError(
                        "AI gateway returned "
                        "no usable response."
                    )

                # ------------------------------------------------
                # Stream text AFTER the successful model call.
                #
                # We intentionally buffer the model call above so
                # transient gateway failures don't leave the UI
                # with half a broken answer.
                # ------------------------------------------------

                if assistant_text:

                    # Send the complete answer as one coherent
                    # logical event. The frontend can render it
                    # continuously without receiving broken fragments.
                    yield sse(
                        "token",
                        {
                            "text":
                                assistant_text
                        },
                    )

                # ------------------------------------------------
                # No tool calls = completed answer.
                # ------------------------------------------------

                if not tool_calls:

                    yield sse(
                        "done",
                        {
                            "round":
                                round_index + 1,
                            "model":
                                current_model,
                        },
                    )

                    logger.info(
                        "AI request completed | "
                        "model=%s | round=%s",
                        current_model,
                        round_index + 1,
                    )

                    return

                # ------------------------------------------------
                # Tool calls
                # ------------------------------------------------

                serialized = []

                for index in sorted(
                    tool_calls
                ):

                    tool_call = (
                        tool_calls[index]
                    )

                    serialized.append(
                        {
                            "id":
                                tool_call[
                                    "id"
                                ]
                                or (
                                    "call_"
                                    + uuid.uuid4().hex
                                ),
                            "type":
                                "function",
                            "function":
                                {
                                    "name":
                                        tool_call[
                                            "name"
                                        ],
                                    "arguments":
                                        tool_call[
                                            "arguments"
                                        ],
                                },
                        }
                    )

                messages.append(
                    {
                        "role":
                            "assistant",
                        "content":
                            assistant_text
                            or None,
                        "tool_calls":
                            serialized,
                    }
                )

                for tool_call in serialized:

                    name = (
                        tool_call[
                            "function"
                        ]["name"]
                    )

                    arguments = (
                        tool_call[
                            "function"
                        ]["arguments"]
                    )

                    yield sse(
                        "tool_start",
                        {
                            "name":
                                name,
                        },
                    )

                    result = await asyncio.to_thread(
                        execute_tool,
                        name,
                        arguments,
                        req.chat_id,
                    )

                    yield sse(
                        "tool_result",
                        {
                            "name":
                                name,
                            "result":
                                result[
                                    :6000
                                ],
                        },
                    )

                    messages.append(
                        {
                            "role":
                                "tool",
                            "tool_call_id":
                                tool_call[
                                    "id"
                                ],
                            "name":
                                name,
                            "content":
                                result,
                        }
                    )

                # Tools succeeded. Continue the agent loop.
                break

            except Exception as exc:

                message = error_text(
                    exc
                )

                logger.error(
                    "AI gateway failure | "
                    "model=%s | "
                    "round=%s | "
                    "attempt=%s/%s | "
                    "%s",
                    current_model,
                    round_index + 1,
                    transient_failures + 1,
                    AI_RETRIES + 1,
                    message,
                )

                # ------------------------------------------------
                # Tool compatibility fallback
                # ------------------------------------------------

                if (
                    tools_enabled
                    and any(
                        marker in str(
                            exc
                        ).lower()
                        for marker in (
                            "tool",
                            "function",
                            "invalid_request_error",
                        )
                    )
                ):

                    tools_enabled = False

                    yield sse(
                        "status",
                        {
                            "message":
                                "This model "
                                "does not accept "
                                "the current tool "
                                "format. Retrying "
                                "without tools…"
                        },
                    )

                    transient_failures = 0

                    continue

                # ------------------------------------------------
                # Temporary gateway failure
                # ------------------------------------------------

                if is_transient_error(
                    exc
                ):

                    if (
                        transient_failures
                        < AI_RETRIES
                    ):

                        wait_seconds = min(
                            2 ** transient_failures,
                            8,
                        )

                        yield sse(
                            "status",
                            {
                                "message":
                                    "The AI gateway "
                                    "is temporarily "
                                    "unavailable. "
                                    f"Retrying "
                                    f"({transient_failures + 1}/"
                                    f"{AI_RETRIES})…"
                            },
                        )

                        await asyncio.sleep(
                            wait_seconds
                        )

                        transient_failures += 1

                        continue

                    # ------------------------------------------------
                    # Try a fallback model.
                    # ------------------------------------------------

                    if (
                        model_index
                        + 1
                        < len(
                            model_attempts
                        )
                    ):

                        model_index += 1

                        fallback_model = (
                            model_attempts[
                                model_index
                            ]
                        )

                        yield sse(
                            "status",
                            {
                                "message":
                                    "The selected "
                                    "model is "
                                    "temporarily "
                                    "unavailable. "
                                    "Trying a "
                                    "fallback model…",
                                "model":
                                    fallback_model,
                            },
                        )

                        break

                # ------------------------------------------------
                # Non-transient error
                # ------------------------------------------------

                yield sse(
                    "error",
                    {
                        "message":
                            (
                                "AI gateway error: "
                                + message
                            )
                    },
                )

                return

        else:

            yield sse(
                "error",
                {
                    "message":
                        "The AI gateway did not "
                        "respond after multiple "
                        "attempts.",
                },
            )

            return

    yield sse(
        "error",
        {
            "message":
                "Agent reached its safety "
                f"limit of "
                f"{MAX_TOOL_ROUNDS} "
                "execution rounds before "
                "completion.",
        },
    )


# ============================================================
# ROUTES
# ============================================================

@app.get(
    "/",
    response_class=HTMLResponse,
)
async def index():

    return Path(
        "static/index.html"
    ).read_text(
        encoding="utf-8"
    )


@app.get(
    "/api/health"
)
async def health():

    return {
        "ok": True,
        "app": APP_NAME,
        "api_configured":
            bool(API_KEY),
        "agent": True,
        "agent_version":
            "4.1.0",
        "tool_rounds":
            MAX_TOOL_ROUNDS,
        "gateway":
            BASE_URL,
        "timestamp":
            int(time.time()),
    }


@app.get(
    "/api/models"
)
async def models():

    if not API_KEY:

        return {
            "models":
                [
                    item[1]
                    for item
                    in FALLBACK_MODELS
                ],
            "source":
                "fallback",
            "api_configured":
                False,
        }

    client = AsyncOpenAI(
        base_url=BASE_URL,
        api_key=API_KEY,
        timeout=30,
        max_retries=0,
    )

    ids = await get_models(
        client
    )

    return {
        "models": ids,
        "source":
            "gateway",
        "api_configured":
            True,
    }


@app.post(
    "/api/chat/stream"
)
async def chat_stream(
    req: ChatRequest
):

    if not req.messages:

        raise HTTPException(
            status_code=400,
            detail=(
                "messages cannot be empty"
            ),
        )

    chat_root(
        req.chat_id
    )

    return StreamingResponse(
        stream_chat(req),
        media_type=(
            "text/event-stream"
        ),
        headers={
            "Cache-Control":
                "no-cache, no-transform",
            "Connection":
                "keep-alive",
            "X-Accel-Buffering":
                "no",
        },
    )


@app.post(
    "/api/upload"
)
async def upload(
    chat_id: str = Form(...),
    files: list[
        UploadFile
    ] = File(...),
):

    root = chat_root(
        chat_id
    )

    saved = []

    for upload_file in files:

        name = Path(
            upload_file.filename
            or "file"
        ).name

        if not name:
            continue

        data = await upload_file.read()

        if (
            len(data)
            > MAX_FILE_BYTES
        ):

            raise HTTPException(
                status_code=413,
                detail=(
                    f"{name} exceeds "
                    "MAX_FILE_BYTES"
                ),
            )

        target = (
            root
            / "uploads"
            / (
                uuid.uuid4().hex[
                    :8
                ]
                + "_"
                + name
            )
        )

        target.write_bytes(
            data
        )

        saved.append(
            {
                "name":
                    name,
                "stored":
                    str(
                        target.relative_to(
                            root
                        )
                    ),
                "size":
                    len(data),
                "type":
                    upload_file.content_type
                    or "application/octet-stream",
            }
        )

    return {
        "files":
            saved
    }


@app.post(
    "/api/download"
)
async def download(
    req: DownloadRequest
):

    root = chat_root(
        req.chat_id
    )

    buffer = io.BytesIO()

    safe_title = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        req.title,
    ).strip(
        "._-"
    ) or "AI_Canvas_Chat"

    with zipfile.ZipFile(
        buffer,
        "w",
        zipfile.ZIP_DEFLATED,
    ) as archive:

        lines = [
            f"# {req.title}",
            "",
            (
                "Exported: "
                + time.strftime(
                    "%Y-%m-%d %H:%M:%S UTC",
                    time.gmtime(),
                )
            ),
            "",
        ]

        for message in req.messages:

            role = str(
                message.get(
                    "role",
                    "unknown",
                )
            ).capitalize()

            content = message.get(
                "content",
                "",
            )

            if isinstance(
                content,
                list,
            ):

                content = json.dumps(
                    content,
                    ensure_ascii=False,
                    indent=2,
                )

            lines += [
                f"## {role}",
                "",
                str(content),
                "",
            ]

        archive.writestr(
            "conversation.md",
            "\n".join(lines),
        )

        for folder in (
            "uploads",
            "outputs",
            "files",
        ):

            base = (
                root / folder
            )

            if base.exists():

                for path in base.rglob("*"):

                    if path.is_file():

                        archive.write(
                            path,
                            path.relative_to(
                                root
                            ).as_posix(),
                        )

        manifest = {
            "title":
                req.title,
            "chat_id":
                req.chat_id,
            "agent_runtime":
                "AI Canvas Agent v4.1",
            "files":
                {},
        }

        for folder in (
            "uploads",
            "outputs",
            "files",
        ):

            base = (
                root / folder
            )

            manifest[
                "files"
            ][folder] = (
                [
                    path.relative_to(
                        root
                    ).as_posix()
                    for path
                    in base.rglob("*")
                    if path.is_file()
                ]
                if base.exists()
                else []
            )

        archive.writestr(
            "manifest.json",
            json.dumps(
                manifest,
                ensure_ascii=False,
                indent=2,
            ),
        )

    return Response(
        buffer.getvalue(),
        media_type=(
            "application/zip"
        ),
        headers={
            "Content-Disposition":
                (
                    'attachment; '
                    f'filename="{safe_title}.zip"'
                )
        },
    )


# ============================================================
# LOCAL ENTRYPOINT
# ============================================================

if __name__ == "__main__":

    import uvicorn

    uvicorn.run(
        app,
        host="0.0.0.0",
        port=int(
            os.getenv(
                "PORT",
                "8000",
            )
        ),
    )

